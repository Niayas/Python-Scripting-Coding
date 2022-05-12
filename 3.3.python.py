
import os
import openpyxl

os.getcwd()
os.chdir(r"C:\Users\niasco6566\Desktop\week 3")
os.listdir()
workBook=openpyxl.load_workbook("example.xlsx")
workSheet=workBook["Sheet1"]
numRows=workSheet.max_row
numCols=workSheet.max_column
letters=[]
for charCode in range(65,91):
    letters.append(chr(charCode))
    
for row in range(1,numRows+1):
    print("|",end="")
    for column in range(numCols):
            print(workSheet[letters[column]+str(row)].value,end="|")
    print("\n")

for row in range(1,numRows+1):
    workSheet['C'+str(row)].value=50
    
for row in range(1,numRows+1):
    print("|",end="")
    for column in range(numCols):
        print(workSheet[letters[column]+str(row)].value,end="|")
    print("\n")
    
workBook.save("newData.xlsx")

workBook=openpyxl.load_workbook("example.xlsx")
workSheet=workBook["Sheet1"]
getData=openpyxl.chart.Reference(workSheet,min_row=1,max_row=numRows,min_col=2,max_col=numCols)
dataRep=openpyxl.chart.Series(getData,title="Fruits")
chart=openpyxl.chart.BarChart()
chart.append(dataRep)
chart.title="Fruits Chart"
workSheet.add_chart(chart,"E1")
workBook.save("newChart.xlsx")
